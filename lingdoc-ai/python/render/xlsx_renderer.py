#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 表格智能填表渲染器 v2
支持 .xlsx / .xls 格式
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


def clean_text(text):
    if not text:
        return ""
    return str(text).strip().replace("：", "").replace(":", "").replace(" ", "").replace("\n", "").replace("\t", "").replace("[", "").replace("]", "").replace("{", "").replace("}", "")


def is_placeholder(text):
    """
    严格判断是否是占位符单元格
    - 空单元格：不是占位符（只是没内容）
    - [xxx] 或 {{xxx}}：是占位符
    - 包含"请填写/输入"等提示文字：是占位符
    """
    if not text:
        return False
    t = str(text).strip()
    if t.startswith("[") and t.endswith("]"):
        return True
    if t.startswith("{{") and t.endswith("}}"):
        return True
    if "请" in t and ("填写" in t or "输入" in t or "不少于" in t or "示例" in t):
        return True
    return False


def get_main_cell(ws, cell):
    """获取合并单元格的主单元格（左上角，可写）"""
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _is_field_name(text, all_field_names):
    """判断文本是否是某个字段名"""
    clean = clean_text(text)
    if not clean:
        return False
    for name in all_field_names:
        if clean == clean_text(name):
            return True
    return False


def fill_excel_document(input_path: str, output_path: str, filled_values: dict):
    wb = load_workbook(input_path)
    filled_count = 0
    unmatched = []
    matched = []
    filled_cells = set()  # (sheet, row, col) 去重
    all_field_names = list(filled_values.keys())

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # === 策略1: 直接匹配 [字段名] 或 {{字段名}} 占位符 ===
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value:
                    continue
                cell_str = str(cell.value).strip()
                for field_name, field_value in filled_values.items():
                    placeholder_bracket = f"[{field_name}]"
                    placeholder_brace = f"{{{{{field_name}}}}}"
                    if placeholder_bracket in cell_str or placeholder_brace in cell_str:
                        main_cell = get_main_cell(ws, cell)
                        key = (sheet_name, main_cell.row, main_cell.column)
                        if key not in filled_cells:
                            main_cell.value = field_value
                            filled_cells.add(key)
                            filled_count += 1
                            matched.append({
                                "fieldName": field_name,
                                "value": field_value,
                                "sheet": sheet_name,
                                "cell": f"{main_cell.column_letter}{main_cell.row}",
                                "matchType": "direct_placeholder"
                            })

        # === 策略2: 字段名匹配 → 找右侧/下方占位符 ===
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value:
                    continue
                cell_text = clean_text(cell.value)
                if not cell_text:
                    continue

                for field_name, field_value in filled_values.items():
                    clean_name = clean_text(field_name)
                    if not clean_name:
                        continue

                    # 严格匹配字段名（标题行）
                    if cell_text == clean_name:
                        target = _find_target_cell(ws, cell, filled_cells, sheet_name, all_field_names)
                        if target:
                            main_cell = get_main_cell(ws, target)
                            key = (sheet_name, main_cell.row, main_cell.column)
                            if key not in filled_cells:
                                main_cell.value = field_value
                                filled_cells.add(key)
                                filled_count += 1
                                matched.append({
                                    "fieldName": field_name,
                                    "value": field_value,
                                    "sheet": sheet_name,
                                    "cell": f"{main_cell.column_letter}{main_cell.row}",
                                    "matchType": "neighbor"
                                })

    wb.save(output_path)

    # 检查未匹配的字段
    matched_names = set(m["fieldName"] for m in matched)
    for field_name in filled_values.keys():
        if field_name not in matched_names:
            unmatched.append({"fieldName": field_name, "reason": "未找到对应占位符"})

    return {
        "success": True,
        "filledCount": filled_count,
        "matchedFields": matched,
        "unmatchedFields": unmatched,
        "outputPath": output_path
    }


def _find_target_cell(ws, field_cell, filled_cells, sheet_name, all_field_names):
    """
    找到字段名对应的值单元格
    策略：优先右侧，其次下方
    排除：
    - 空单元格（不是占位符，只是没内容）
    - 其他字段名标签（避免串行）
    """
    r, c = field_cell.row, field_cell.column

    candidates = []

    # 右侧单元格
    if c + 1 <= ws.max_column:
        right = ws.cell(row=r, column=c + 1)
        key = (sheet_name, right.row, right.column)
        if key not in filled_cells and is_placeholder(right.value) and not _is_field_name(right.value, all_field_names):
            candidates.append((right, 1))  # 优先级1

    # 下方单元格
    if r + 1 <= ws.max_row:
        down = ws.cell(row=r + 1, column=c)
        key = (sheet_name, down.row, down.column)
        if key not in filled_cells and is_placeholder(down.value) and not _is_field_name(down.value, all_field_names):
            candidates.append((down, 2))  # 优先级2

    # 隔一列右侧
    if c + 2 <= ws.max_column:
        right2 = ws.cell(row=r, column=c + 2)
        key = (sheet_name, right2.row, right2.column)
        if key not in filled_cells and is_placeholder(right2.value) and not _is_field_name(right2.value, all_field_names):
            candidates.append((right2, 3))  # 优先级3

    if candidates:
        candidates.sort(key=lambda x: x[1])
        return candidates[0][0]
    return None


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python xlsx_renderer.py <input.xlsx> <output.xlsx> '{\"姓名\":\"张三\"}'")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]
    filled_values = json.loads(sys.argv[3])

    result = fill_excel_document(input_path, output_path, filled_values)
    print(json.dumps(result, ensure_ascii=False, indent=2))
